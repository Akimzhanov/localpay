from django.contrib import admin 
from datetime import datetime
from django.contrib.auth.hashers import make_password
from .models import User_mon, Pays, Comment
from rangefilter.filters import DateRangeFilter, DateTimeRangeFilter, NumericRangeFilter
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Sum,FloatField
from django.db.models.functions import Cast
from django.db.models import Q
from django.contrib import messages
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font





@admin.register(User_mon)
class UserAdmin(admin.ModelAdmin):
    exclude = ['date_reg' , 'last_login']
    list_display = ['id','name', 'surname', 'login','date_reg', 'access', 'balance', 'avail_balance','region']
    search_fields = ['name','surname']
    readonly_fields = ('balance','avail_balance')


    def save_model(self, request, obj, form, change):
        # Получаем список измененных полей
        changed_fields = form.changed_data

        if not obj.pk:
            # Создаем новый экземпляр User_mon
            obj = form.save(commit=False)

            # Хешируем пароль
            obj.password = make_password(obj.password)

            # Добавляем дату регистрации
            obj.date_reg = datetime.now()

            # Сохраняем объект
            obj.save()

            # Возвращаем обновленный объект
            form.save_m2m()
            return obj


        else:
            # Обновляем только измененные поля
            for field_name in changed_fields:

                if field_name == 'password':
                    password = obj.password
                    password2 = make_password(password)
                    setattr(obj, field_name, password2)
                    obj.save(update_fields=['password'] + changed_fields)
                else:
                    setattr(obj, field_name, form.cleaned_data.get(field_name))
                    obj.save(update_fields=['login', 'surname', 'name', 'access','is_active'] + changed_fields)
                

            #Пополнение счета с бухгалтерии
            if obj.refill > 0:
                obj.comment = 'Пополнение'
                old_balance = obj.balance
                new_balance = obj.balance + obj.write_off
                obj.balance = new_balance
        
                time = datetime.now()
                time = str(time)[:-4]
                b = 'Пополнение с бухгалтерии'
                ls = '*********'
                check_pay = Pays.objects.create(user=obj,date_payment=time,accept_payment=time,ls_abon=ls,money=obj.refill, status_payment=b)
                check_pay.save()    



                old_balance = obj.balance      #старый баланс
                balance = old_balance + obj.refill   #общий баланс
                obj.balance = balance     #присваиваем старый баланс 
                old_avail_balance = obj.avail_balance   #старый потраченый баланс до изменения
                avail_balance = old_avail_balance + obj.write_off   # общ потраченый баланс
                obj.avail_balance = avail_balance   # присваиваем старый баланс 




                Comment.objects.create(
                    user2=obj,
                    type_pay=obj.comment,
                    old_balance=old_balance,
                    new_balance=obj.refill,
                    mont_balance=balance,
                    old_avail_balance=old_avail_balance,
                    new_avail_balance=obj.write_off,
                    mont_avail_balance=avail_balance
                )

                obj.refill = 0
                obj.write_off = 0
                obj.comment = ''
                # Сохраняем объект
                obj.save(update_fields=['balance', 'avail_balance', 'comment', 'password'] + changed_fields)

            #Списание счета с бухгалтерии
            elif obj.write_off > 0:
                obj.comment = 'Списание'
                old_avail_balance = obj.avail_balance
                new_balance = obj.balance + obj.write_off

        
                time = datetime.now()
                time = str(time)[:-4]
                b = 'Списание с бухгалтерии'
                ls = '*********'
                check_pay = Pays.objects.create(user=obj,date_payment=time,accept_payment=time,ls_abon=ls,money=obj.write_off, status_payment=b)
                check_pay.save()    



                old_balance = obj.balance      #старый баланс
                balance = old_balance + obj.write_off   #общий баланс
                obj.balance = balance     #присваиваем старый баланс 
                old_avail_balance = obj.avail_balance   #старый потраченый баланс до изменения
                avail_balance = old_avail_balance + obj.write_off   # общ потраченый баланс
                obj.avail_balance = avail_balance   # присваиваем старый баланс 




                Comment.objects.create(
                    user2=obj,
                    type_pay=obj.comment,
                    old_balance=old_balance,
                    new_balance=obj.write_off,
                    mont_balance=balance,
                    old_avail_balance=old_avail_balance,
                    new_avail_balance=obj.write_off,
                    mont_avail_balance=avail_balance
                )

                obj.refill = 0
                obj.write_off = 0
                obj.comment = ''

                # Сохраняем объект
                obj.save(update_fields=['balance', 'avail_balance', 'comment', 'password'] + changed_fields)

def make_published(PayAdmin, request, queryset):
    response = queryset.values()
    money = 0
    for i in response:
        if i['status_payment'] == 'Выполнен' and i['annulment'] == False:
            c = i['money']
            test2 = c[:-2]
            money +=float(test2) if test2 else 0

    messages.info(request, f"В общем было оплачено: {money} сом")


def export_to_excel_user(User_mon, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=User_.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'User_mon'
    # write header row
    header = ['Имя', 'Фамилия', 'Логин', 'Баланс', 'Сумма оплат']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    # write data rows
    for row_num, obj in enumerate(queryset, 1):
        row = [
            obj.name,
            obj.surname,
            obj.login,
            obj.balance, 
            obj.avail_balance]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
 








def export_to_excel_01(PayAdmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Pays_.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pays'
    # write header row
    header = ['Номер платежа', 'Время проведения платежа', 'Лицевой счет', 'Деньги', 'Статус платежа', 'Пользователь', 'Аннуляция', 'Комментарий']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    # write data rows
    for row_num, obj in enumerate(queryset, 1):
        row = [
            obj.number_payment,
            obj.accept_payment,
            obj.ls_abon,
            obj.money, 
            obj.status_payment,
            str(obj.user),
            obj.annulment,
            obj.comment
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
 




@admin.register(Pays)
class PayAdmin(admin.ModelAdmin):
    exclude = ['number_payment']
    list_display = ['date_payment', 'ls_abon', 'money', 'status_payment', 'user', 'annulment']
    list_filter = [('date_payment', DateRangeFilter)]
    search_fields = ['user__name', 'user__surname', 'status_payment','ls_abon']
    list_per_page = 600 

    def export_to_excel(self, request, queryset):
        start_date = request.GET.get('date_payment__range__gte')
        end_date = request.GET.get('date_payment__range__lte')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Pays_{start_date} - {end_date}.xlsx'
    
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Pays'
    
        users = set(queryset.values_list('user', flat=True))  # Получаем уникальных пользователей из queryset
    
        row_num = 1  # Стартовая позиция строки
        total_money_all_users = 0  # Общая сумма для всех пользователей
    
        for user_id in users:
            user = User_mon.objects.get(id=user_id)  # Получаем объект пользователя по user_id
    
            # Добавляем отступы между пользователями
            row_num += 3
    
            # Заголовок таблицы для текущего пользователя
            row_num += 1
            cell = worksheet.cell(row=row_num, column=1) 
            cell.value = f'Платежи пользователя: {user}'
            cell.font = Font(bold=True)
    
            # Выравнивание заголовка таблицы по центру
            max_column = worksheet.max_column
            cell_start = worksheet.cell(row=row_num, column=1)
            cell_end = worksheet.cell(row=row_num, column=max_column)
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_column)
            cell_start.alignment = Alignment(horizontal='center')
    
            # Заголовки столбцов
            header = ['Номер платежа', 'Время проведения платежа', 'Лицевой счет', 'Деньги', 'Статус платежа', 'Пользователь', 'Аннуляция', 'Комментарий']
            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True)
    
            # Платежи текущего пользователя
            user_payments = queryset.filter(user=user)
            total_money = 0  # Общая сумма поля "money" для текущего пользователя
            for payment_num, payment in enumerate(user_payments, 2):
                row_num += 2
                row = [
                    payment.number_payment,
                    payment.accept_payment,
                    payment.ls_abon,
                    payment.money,
                    payment.status_payment,
                    str(payment.user),
                    payment.annulment,
                    payment.comment 
                ]
                if payment.status_payment =='Выполнен' and not payment.annulment:
                    total_money += float(payment.money)  # Обновляем общую сумму
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
    
            # Добавляем общую сумму на следующей строке под столбцом "Деньги"
            row_num += 1
            total_money_cell = worksheet.cell(row=row_num, column=header.index('Деньги') + 1)
            total_money_cell.value = total_money
    
            # Добавляем сумму текущего пользователя к общей сумме всех пользователей
            total_money_all_users += total_money
    
        # Добавляем общую сумму всех пользователей в последней строке
        row_num += 2
        total_money_all_users_cell = worksheet.cell(row=row_num, column=header.index('Деньги') + 1)
        total_money_all_users_cell.value = f'Общая сумма : {total_money_all_users}'
        total_money_all_users_cell.font = Font(bold=True)
    
        # Устанавливаем ширину столбцов на основе содержимого
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
        workbook.save(response)
        return response
    
    export_to_excel.short_description = 'Выгрузить в Excel'
    
    actions = [export_to_excel]




    # Работа ануллироания и добавление аннулированных денег обратно на счет монтажника
    def save_model(self, request, obj, form, change):
        if obj.annulment == True:
            user = obj.user
            money = obj.money
            x = money[:-2]
            y = int(x)
            test = user.balance + y
            user.balance = test
            test2 = user.avail_balance + y
            user.avail_balance = test2
            anul = 'Аннулирован'

            old_balance = user.balance - y
            old_avail_balance = user.avail_balance - y
            a = Comment.objects.create(user2=user, text=obj.comment,type_pay=anul,old_balance=old_balance, new_balance=y, mont_balance=test, old_avail_balance = old_avail_balance,new_avail_balance=y,mont_avail_balance=test2)
            a.save()
            user.save(update_fields=['balance', 'avail_balance'])

        obj.save(update_fields=['annulment'])





@admin.register(Comment)
class CommentAdmin(admin.ModelAdmin):
    list_display = ['user2', 'text','type_pay','old_balance','new_balance','mont_balance','old_avail_balance','new_avail_balance','mont_avail_balance','created_at']
    search_fields = ['user2__name', 'user2__surname']



admin.site.add_action(make_published, 'Общая сумма')
admin.site.add_action(export_to_excel_01, 'Выгрузить в Excel(старая версия)')
admin.site.add_action(export_to_excel_user, 'Выгрузка пользователей')



